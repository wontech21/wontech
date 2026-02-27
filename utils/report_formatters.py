"""
Report format generators for WONTECH reporting engine.

Each generator follows the signature:
    generate_X(headers, rows, report_meta) -> (bytes, mime_type, extension)

Where report_meta is a dict with keys: key, name, category, description.
"""

import csv
import io
from datetime import datetime


def generate_csv(headers, rows, report_meta):
    """Generate a CSV file from report data."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    csv_bytes = buf.getvalue().encode('utf-8')
    return (csv_bytes, 'text/csv', 'csv')


def generate_xlsx(headers, rows, report_meta):
    """Generate a branded XLSX file from report data."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    # Excel sheet names are limited to 31 characters
    ws.title = report_meta['name'][:31]

    # Row 1: Report title
    ws.cell(row=1, column=1, value=report_meta['name']).font = Font(
        bold=True, size=14
    )

    # Row 2: Generated date
    ws.cell(row=2, column=1, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # Row 3: empty (spacer)

    # Row 4: Headers with WONTECH branding
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Row 5+: Data rows with alternating colors
    alt_fill = PatternFill(start_color='f9fafb', end_color='f9fafb', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for row_idx, row in enumerate(rows):
        excel_row = row_idx + 5
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = fill

    # Auto-width columns (cap at 50)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    return (xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx')


def generate_pdf(headers, rows, report_meta):
    """Generate a branded PDF file from report data."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise ImportError(
            "reportlab is required for PDF export. "
            "Install it with: pip install reportlab"
        )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)

    brand_color = colors.HexColor('#667eea')
    alt_color = colors.HexColor('#f9fafb')
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Title'], fontSize=20, spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle', parent=styles['Normal'], fontSize=10,
        textColor=colors.grey, spaceAfter=12
    )

    elements = []
    elements.append(Paragraph(report_meta['name'], title_style))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style
    ))
    elements.append(Spacer(1, 0.25 * inch))

    # Build table data
    table_data = [headers] + [list(row) for row in rows]
    table = Table(table_data, repeatRows=1)

    # Table styling
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), brand_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
    ]
    # Alternating row colors
    for i in range(1, len(table_data)):
        if i % 2 == 1:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), alt_color))

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.grey)
        canvas.drawString(inch, 0.5 * inch, "Generated by WONTECH")
        canvas.drawRightString(
            letter[0] - inch, 0.5 * inch, f"Page {doc.page}"
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    pdf_bytes = buf.getvalue()
    return (pdf_bytes, 'application/pdf', 'pdf')
